"""bytes / 貼り付けテキスト → SheetGrid。

判定エンジンで唯一 I/O に近い層（xlsx のときだけ openpyxl を遅延 import する）。
ここも例外は投げず、迷ったら警告を残して最善のグリッドを返す。
"""

from __future__ import annotations

import csv
import io
import re
import zipfile
from datetime import date, datetime

from .contracts import SheetGrid, SourceMeta

# 区切り文字の候補（優先順）。この順でスコアが同点なら先のものを採る。
DELIMITER_CANDIDATES = (",", "\t", ";", "|")
# 区切り文字と認めるスコアの下限。下回ったら複数空白区切りを試す。
MIN_DELIMITER_SCORE = 0.6
# 1 ファイルあたりの行数上限（これを超えたら切って警告）
MAX_ROWS = 50_000


# ----------------------------------------------------------------------
# 種別の判定
# ----------------------------------------------------------------------


def sniff_kind(data: bytes) -> str:
    """csv | xlsx | xls | pdf | unknown。"""
    if not data:
        return "unknown"
    if data[:4] == b"PK\x03\x04":
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as zf:
                names = zf.namelist()
            if any(n.startswith("xl/") for n in names):
                return "xlsx"
        except zipfile.BadZipFile:
            return "unknown"
        return "unknown"          # docx や単なる zip
    if data[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "xls"              # 旧 Excel。読まずに案内する
    if data[:4] == b"%PDF":
        return "pdf"
    return "csv"


# ----------------------------------------------------------------------
# 文字コード
# ----------------------------------------------------------------------

_BOMS = (
    (b"\xef\xbb\xbf", "utf-8-sig"),
    (b"\xff\xfe\x00\x00", "utf-32-le"),
    (b"\x00\x00\xfe\xff", "utf-32-be"),
    (b"\xff\xfe", "utf-16-le"),
    (b"\xfe\xff", "utf-16-be"),
)

_ENCODING_LADDER = ("utf-8", "cp932", "euc-jp", "iso-2022-jp")

_JP_RANGES = (
    (0x3040, 0x30FF),   # かな
    (0x4E00, 0x9FFF),   # 漢字
    (0xFF00, 0xFFEF),   # 全角英数・半角カナ
)


def _jp_text_score(text: str) -> float:
    """日本語テキストとしての尤もらしさ（文字化け検出用）。"""
    if not text:
        return 0.0
    good = bad = 0
    for ch in text[:5000]:
        cp = ord(ch)
        if ch in "\r\n\t":
            continue
        if cp == 0xFFFD:
            bad += 3
            continue
        if 0x20 <= cp < 0x7F:
            good += 1
        elif any(lo <= cp <= hi for lo, hi in _JP_RANGES):
            good += 1
        elif 0x80 <= cp <= 0xFF:
            # Latin-1 補助が続くのは cp932 を utf-8 と誤読したときの典型
            bad += 1
        else:
            good += 1
    total = good + bad
    return (good / total) if total else 0.0


def decode_text(data: bytes) -> tuple[str, str, list[str]]:
    """bytes → (テキスト, 使った文字コード, 警告)。"""
    warnings: list[str] = []
    if not data:
        return ("", "utf-8", warnings)

    for bom, enc in _BOMS:
        if data.startswith(bom):
            try:
                return (data.decode(enc), enc, warnings)
            except UnicodeDecodeError:
                break

    decoded: dict[str, str] = {}
    for enc in _ENCODING_LADDER:
        try:
            decoded[enc] = data.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue

    if not decoded:
        warnings.append("文字コードを判別できませんでした。読めない文字は置き換えています")
        return (data.decode("utf-8", errors="replace"), "utf-8", warnings)

    # UTF-8 は自己検証的なので、通ればほぼ UTF-8。ただし化けの兆候があれば cp932 を見る。
    if "utf-8" in decoded:
        text = decoded["utf-8"]
        score = _jp_text_score(text)
        if score < 0.9 and "cp932" in decoded:
            alt = decoded["cp932"]
            if _jp_text_score(alt) > score:
                warnings.append("文字コードを cp932 と判定しました（UTF-8 として読むと文字化けするため）")
                return (alt, "cp932", warnings)
        return (text, "utf-8", warnings)

    for enc in _ENCODING_LADDER[1:]:
        if enc in decoded:
            if enc != "cp932":
                warnings.append(f"文字コードを {enc} と判定しました")
            return (decoded[enc], enc, warnings)

    warnings.append("文字コードを判別できませんでした。読めない文字は置き換えています")
    return (data.decode("utf-8", errors="replace"), "utf-8", warnings)


# ----------------------------------------------------------------------
# 区切り文字
# ----------------------------------------------------------------------


def _rows_for(text: str, delimiter: str) -> list[list[str]]:
    """csv.reader で読む（銘柄名に含まれるカンマを引用符ごと扱うため必須）。"""
    try:
        return [row for row in csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)]
    except csv.Error:
        return []


def _consistency_score(rows: list[list[str]]) -> tuple[float, int]:
    """(スコア, 最頻列数)。前文が混ざっていても最長の等列数区間で拾えるようにする。"""
    counts = [len(r) for r in rows if any(c.strip() for c in r)]
    if not counts:
        return (0.0, 0)
    freq: dict[int, int] = {}
    for c in counts:
        freq[c] = freq.get(c, 0) + 1
    mode = max(freq, key=lambda k: (freq[k], k))
    if mode < 2:
        return (0.0, mode)

    overall = freq[mode] / len(counts)

    # 最長連続区間でも測る。12 行のうち 4 行が前文、というときは全体比だと沈む。
    best_run = run = 0
    for c in counts:
        run = run + 1 if c == mode else 0
        best_run = max(best_run, run)
    run_ratio = best_run / len(counts)

    consistency = max(overall, run_ratio)
    return (consistency * (1 - 1 / mode), mode)


def sniff_delimiter(text: str) -> tuple[str | None, str, float]:
    """(区切り文字, モード, スコア)。モードは char | multispace | single_column。"""
    if not text.strip():
        return (None, "single_column", 0.0)

    best: tuple[float, int, str | None] = (0.0, 0, None)
    for cand in DELIMITER_CANDIDATES:
        rows = _rows_for(text, cand)
        score, mode = _consistency_score(rows)
        if score > best[0] or (score == best[0] and mode > best[1]):
            best = (score, mode, cand)

    if best[0] >= MIN_DELIMITER_SCORE and best[2] is not None:
        return (best[2], "char", best[0])

    # 画面からの貼り付け: タブか連続空白
    rows = _split_multispace(text)
    score, _mode = _consistency_score(rows)
    if score >= MIN_DELIMITER_SCORE:
        return (None, "multispace", score)

    if best[2] is not None and best[0] > 0:
        return (best[2], "char", best[0])
    return (None, "single_column", 0.0)


def _looks_like_block_paste(text: str) -> bool:
    """1件が複数行に分かれた貼り付けか。

    日付で始まる行が飛び飛びに現れ、その間に区切りの無い行が続く形。
    表としては読めないが、原因を名指しできれば利用者は次の手を打てる。
    """
    from .shapes import parse_date

    lines = [ln for ln in text.splitlines() if ln.strip()]
    if len(lines) < 6:
        return False
    starts = sum(
        1 for ln in lines if parse_date(re.split(r"[\t,;|]", ln, maxsplit=1)[0])
    )
    # 日付で始まる行が数本あり、かつ大半の行がそうでない（＝間に続きがある）
    return starts >= 3 and starts * 2 < len(lines)


_MULTISPACE_RE = re.compile(r"\t|[ 　]{2,}")


def _split_multispace(text: str) -> list[list[str]]:
    out: list[list[str]] = []
    for line in text.splitlines():
        if not line.strip():
            out.append([])
            continue
        out.append([c.strip() for c in _MULTISPACE_RE.split(line.strip())])
    return out


# ----------------------------------------------------------------------
# xlsx
# ----------------------------------------------------------------------


def _cell_text(value: object) -> tuple[str, str]:
    """xlsx のセル値 → (表示テキスト, 型)。"""
    if value is None:
        return ("", "")
    if isinstance(value, datetime):
        return (value.date().isoformat(), "date")
    if isinstance(value, date):
        return (value.isoformat(), "date")
    if isinstance(value, bool):
        return (str(value), "text")
    if isinstance(value, (int, float)):
        text = repr(value) if isinstance(value, float) else str(value)
        if text.endswith(".0"):
            text = text[:-2]
        return (text, "number")
    return (str(value), "text")


def _trim(rows: list[list[tuple[str, str]]]) -> list[list[tuple[str, str]]]:
    while rows and not any(t for t, _ in rows[-1]):
        rows.pop()
    width = 0
    for row in rows:
        for i, (text, _) in enumerate(row):
            if text:
                width = max(width, i + 1)
    return [row[:width] for row in rows]


def load_xlsx(data: bytes, sheet: str | None = None) -> tuple[SheetGrid, list[str]]:
    warnings: list[str] = []
    try:
        import openpyxl                       # 遅延 import（未導入でも他の経路は動く）
    except ImportError:
        warnings.append(
            "Excel ファイルを読むには openpyxl が必要です。CSV で保存し直すか、"
            'pip install -e ".[dev]" で再インストールしてください'
        )
        return (SheetGrid(meta=SourceMeta(kind="xlsx", warnings=tuple(warnings))), warnings)

    def _read(data_only: bool):
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=data_only)
        try:
            names = list(wb.sheetnames)
            picked = sheet if sheet in names else _best_sheet(wb, names)
            ws = wb[picked]
            rows: list[list[tuple[str, str]]] = []
            for i, raw in enumerate(ws.iter_rows(values_only=True)):
                if i >= MAX_ROWS:
                    warnings.append(f"{MAX_ROWS} 行を超えたため以降を読み飛ばしました")
                    break
                rows.append([_cell_text(v) for v in raw])
            return (_trim(rows), picked, names)
        finally:
            wb.close()

    try:
        rows, picked, names = _read(True)
        # data_only=True は計算済みの値を返す。計算エンジンの無いツールが書いた
        # ブックでは全部 None になるので、そのときだけ数式のまま読み直す。
        if rows and not any(t for row in rows for t, _ in row):
            warnings.append("計算結果が保存されていないため、数式のまま読み込みました")
            rows, picked, names = _read(False)
    except Exception as e:  # noqa: BLE001  壊れたブックは利用者側の入力誤り
        warnings.append(f"Excel ファイルを読み込めませんでした（{type(e).__name__}）")
        return (SheetGrid(meta=SourceMeta(kind="xlsx", warnings=tuple(warnings))), warnings)

    grid = SheetGrid(
        rows=tuple(tuple(t for t, _ in row) for row in rows),
        types=tuple(tuple(ty for _, ty in row) for row in rows),
        meta=SourceMeta(
            kind="xlsx",
            sheet_name=picked,
            sheet_names=tuple(names),
            warnings=tuple(warnings),
        ),
    )
    return (grid, warnings)


def _best_sheet(wb, names: list[str]) -> str:
    """データ行が最も多いシートを既定にする。"""
    best, best_score = names[0], -1.0
    for name in names:
        ws = wb[name]
        rows = cols = 0
        for i, raw in enumerate(ws.iter_rows(values_only=True)):
            if i >= 200:
                break
            filled = sum(1 for v in raw if v is not None and str(v).strip())
            if filled >= 3:
                rows += 1
                cols = max(cols, filled)
        score = rows * max(cols, 1)
        if score > best_score:
            best, best_score = name, score
    return best


# ----------------------------------------------------------------------
# 入口
# ----------------------------------------------------------------------


def load_grid(
    data: bytes,
    *,
    filename: str | None = None,
    sheet: str | None = None,
) -> SheetGrid:
    """bytes → SheetGrid。判別できない入力でも例外は投げず、警告を積む。"""
    kind = sniff_kind(data)
    if kind == "xls":
        grid, _ = (
            SheetGrid(
                meta=SourceMeta(
                    kind="xls",
                    filename=filename,
                    warnings=(
                        "旧 Excel 形式(.xls)は読み取れません。.xlsx か CSV で保存し直してください",
                    ),
                )
            ),
            None,
        )
        return grid
    if kind == "pdf":
        return SheetGrid(
            meta=SourceMeta(
                kind="pdf",
                filename=filename,
                warnings=("PDF が渡されました。マネーフォワードMEの資産内訳PDFは「PDF取込」から取り込んでください",),
            )
        )
    if kind == "xlsx":
        grid, _ = load_xlsx(data, sheet=sheet)
        return SheetGrid(rows=grid.rows, types=grid.types,
                         meta=SourceMeta(**{**grid.meta.__dict__, "filename": filename}))

    text, encoding, warnings = decode_text(data)
    return load_text_grid(text, encoding=encoding, filename=filename, warnings=warnings)


def load_text_grid(
    text: str,
    *,
    encoding: str | None = None,
    filename: str | None = None,
    kind: str = "csv",
    warnings: list[str] | None = None,
) -> SheetGrid:
    """テキスト（CSV 本文・貼り付け）→ SheetGrid。"""
    warns = list(warnings or [])
    delimiter, mode, _score = sniff_delimiter(text)

    if mode == "multispace":
        rows = _split_multispace(text)
    elif delimiter is not None:
        rows = _rows_for(text, delimiter)
    else:
        # 「読めなかった」で終わらせず、読めない理由が分かるなら言う。
        # 画面からコピーした表は 1 件の取引が複数行に分かれることがあり
        # （銘柄名がセル内で改行され、口座区分・数量・単価が縦に並ぶ）、
        # 行×列として読めない。まだ対応していないので、その旨を返す。
        if _looks_like_block_paste(text):
            warns.append(
                "1件の取引が複数行に分かれた貼り付けのようです。この形はまだ"
                "読み取れません。CSV でダウンロードできないか確認してください"
            )
        else:
            warns.append("区切り文字を判別できませんでした。1列として読み込みます")
        rows = [[line] for line in text.splitlines()]

    if len(rows) > MAX_ROWS:
        warns.append(f"{MAX_ROWS} 行を超えたため以降を読み飛ばしました")
        rows = rows[:MAX_ROWS]

    cleaned = [tuple(c.strip() if isinstance(c, str) else "" for c in row) for row in rows]
    while cleaned and not any(cleaned[-1]):
        cleaned.pop()

    return SheetGrid(
        rows=tuple(cleaned),
        types=None,
        meta=SourceMeta(
            kind=kind,
            encoding=encoding,
            delimiter=delimiter,
            delimiter_mode=mode,
            filename=filename,
            warnings=tuple(warns),
        ),
    )


def load_pasted(text: str, *, filename: str | None = None) -> SheetGrid:
    """画面からコピーした表をそのまま解析する。"""
    return load_text_grid(text, encoding=None, filename=filename, kind="paste")
